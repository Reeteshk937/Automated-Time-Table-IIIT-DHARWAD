import openpyxl
import os

OUTPUT_DIR = 'output_timetables'
HTML_DIR = '.'

SEMESTERS = [
    ('sem1_timetable.xlsx', 'Semester 1'),
    ('sem3_timetable.xlsx', 'Semester 3'),
    ('sem5_timetable.xlsx', 'Semester 5'),
    ('sem7_timetable.xlsx', 'Semester 7'),
]

def excel_to_html_table(ws):
    html = '<table>\n<thead>\n<tr>'
    for cell in ws[1]:
        html += f'<th>{cell.value}</th>'
    html += '</tr>\n</thead>\n<tbody>'
    for row in ws.iter_rows(min_row=2, values_only=True):
        html += '<tr>' + ''.join(f'<td>{cell if cell is not None else ""}</td>' for cell in row) + '</tr>'
    html += '</tbody>\n</table>'
    return html

def generate_html_files():
    links = []
    for filename, sem_name in SEMESTERS:
        path = os.path.join(OUTPUT_DIR, filename)
        wb = openpyxl.load_workbook(path)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            html_table = excel_to_html_table(ws)
            html_content = f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{sem_name} - {sheet}</title>
<link href="https://fonts.googleapis.com/css?family=Roboto:400,700&display=swap" rel="stylesheet">
<style>
body {{ font-family: 'Roboto', sans-serif; background: #f0f2f8; margin: 0; padding: 0; }}
.container {{ background: #fff; margin: 40px auto; padding: 32px; border-radius: 16px; box-shadow: 0 8px 32px rgba(0,0,0,0.1); max-width: 1000px; }}
h1 {{ color: #2d3a4b; text-align: center; margin-bottom: 24px; }}
table {{ width: 100%; border-collapse: collapse; margin-top: 16px; }}
th, td {{ border: 1px solid #bfc9d1; padding: 12px 8px; text-align: center; }}
th {{ background: #4f8cff; color: #fff; }}
tr:nth-child(even) {{ background: #f0f4fa; }}
tr:hover {{ background: #e3eaff; }}
.button-group {{ text-align: center; margin-bottom: 16px; }}
button, a.button-link {{
  display: inline-block;
  padding: 10px 20px;
  margin: 5px;
  background: #4f8cff;
  color: #fff;
  text-decoration: none;
  border: none;
  border-radius: 6px;
  cursor: pointer;
  font-size: 16px;
}}
button:hover, a.button-link:hover {{ background: #3b6cd1; }}
</style>
</head>
<body>
<div class="container">
<div class="button-group">
<a href="index.html" class="button-link">&#8592; Back</a>
<button onclick="window.print()">🖨️ Print Timetable</button>
<a href="../{OUTPUT_DIR}/{filename}" class="button-link" download>⬇️ Download Excel</a>
</div>
<h1>{sem_name} - {sheet}</h1>
{html_table}
</div>
</body>
</html>'''
            out_name = f"{filename.replace('.xlsx', '')}_{sheet.replace(' ', '_')}.html"
            links.append((sem_name, sheet, out_name))
            with open(os.path.join(HTML_DIR, out_name), 'w', encoding='utf-8') as f:
                f.write(html_content)

    # --- Generate improved index.html ---
    sem_data = {}
    for sem_name, sheet, out_name in links:
        sem_data.setdefault(sem_name, []).append((sheet, out_name))

    index_html = '''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Timetable Selector</title>
<link href="https://fonts.googleapis.com/css?family=Roboto:400,700&display=swap" rel="stylesheet">
<style>
body { font-family: 'Roboto', sans-serif; background: #eef2f7; margin: 0; padding: 0; text-align: center; }
.container { background: #fff; margin: 40px auto; padding: 32px; border-radius: 16px; box-shadow: 0 8px 32px rgba(0,0,0,0.1); max-width: 900px; }
h1 { color: #2d3a4b; margin-bottom: 24px; }
select { padding: 12px 16px; margin: 10px; border-radius: 6px; border: 1px solid #ccc; font-size: 16px; min-width: 220px; }
iframe { width: 100%; height: 650px; border: 2px solid #ccc; border-radius: 12px; margin-top: 25px; background: white; }
</style>
</head>
<body>
<div class="container">
<h1>📘 Timetable Selector</h1>
<label for="semester">Select Semester:</label>
<select id="semester" onchange="updateSections()">
  <option value="">-- Select Semester --</option>
</select>
<label for="section">Select Section:</label>
<select id="section" onchange="showTimetable()">
  <option value="">-- Select Section --</option>
</select>
<iframe id="timetableFrame" src=""></iframe>
</div>
<script>
const semData = {
'''
    for sem_name, sheets in sem_data.items():
        index_html += f'  "{sem_name}": {{\n'
        for sheet, out_name in sheets:
            index_html += f'    "{sheet}": "{out_name}",\n'
        index_html += '  },\n'
    index_html += '''};

const semesterSelect = document.getElementById('semester');
const sectionSelect = document.getElementById('section');
const frame = document.getElementById('timetableFrame');

Object.keys(semData).forEach(sem => {
  const opt = document.createElement('option');
  opt.value = sem;
  opt.textContent = sem;
  semesterSelect.appendChild(opt);
});

function updateSections() {
  const selectedSem = semesterSelect.value;
  sectionSelect.innerHTML = '<option value="">-- Select Section --</option>';
  if (selectedSem && semData[selectedSem]) {
    Object.keys(semData[selectedSem]).forEach(section => {
      const opt = document.createElement('option');
      opt.value = section;
      opt.textContent = section;
      sectionSelect.appendChild(opt);
    });
  }
  frame.src = "";
}

function showTimetable() {
  const sem = semesterSelect.value;
  const sec = sectionSelect.value;
  if (sem && sec && semData[sem][sec]) {
    frame.src = semData[sem][sec];
  }
}
</script>
</body>
</html>'''
    with open(os.path.join(HTML_DIR, 'index.html'), 'w', encoding='utf-8') as f:
        f.write(index_html)
