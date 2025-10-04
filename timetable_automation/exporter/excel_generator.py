import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font

def export_to_excel(assignments, slots_csv="data_templates/slot_mapping_template.csv", output_file="Automated_Timetable.xlsx"):
    """Export timetable to Excel"""
    slots = pd.read_csv(slots_csv).set_index("Slot ID").to_dict("index")

    wb = Workbook()
    ws = wb.active
    ws.title = "Timetable"

    # Header row
    ws.cell(row=1, column=1, value="Day/Time")
    for i, slot in enumerate(slots.keys(), start=2):
        ws.cell(row=1, column=i, value=slot)
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 20

    # Data
    row_num = 2
    for a in assignments:
        ws.cell(row=row_num, column=1, value=a["Course Code"])
        ws.cell(row=row_num, column=2, value=a["Faculty"])
        ws.cell(row=row_num, column=3, value=a["Slot"])
        ws.cell(row=row_num, column=4, value=a["Room"])
        row_num += 1

    # Style
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(size=11)

    wb.save(output_file)
    print(f"✅ Timetable saved as {output_file}")
